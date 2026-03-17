"""
Módulo de exportação/sincronização do CRM para planilha Excel (.xlsx).
Também suporta CSV como fallback.
"""
from pathlib import Path
from datetime import datetime
from typing import Optional
from utils.logger import get_logger

log = get_logger(__name__)

# Colunas do CRM na ordem que aparecerão na planilha
COLUNAS = [
    "id",
    "url_perfil",
    "nome",
    "cargo",
    "empresa",
    "url_empresa",
    "tamanho_empresa",
    "tem_atividade_recente",
    "score_icp",
    "status",
    "notas",
    "mensagem_conexao",
    "mensagem_dm1",
    "data_descoberta",
    "data_ultima_acao",
    "data_conexao",
    "created_at",
]

# Headers legíveis para a planilha
HEADERS = [
    "ID",
    "URL do Perfil",
    "Nome",
    "Cargo",
    "Empresa",
    "URL da Empresa",
    "Tamanho da Empresa",
    "Atividade Recente",
    "Score ICP",
    "Status",
    "Notas",
    "Mensagem de Conexão",
    "Mensagem DM1",
    "Data de Descoberta",
    "Última Ação",
    "Data de Conexão",
    "Criado em",
]


def _lead_to_row(lead) -> list:
    """Converte um registro SQLite Row para lista de valores."""
    return [lead[col] if lead[col] is not None else "" for col in COLUNAS]


def export_to_excel(leads: list, output_path: Optional[Path] = None) -> Path:
    """
    Exporta todos os leads para um arquivo Excel (.xlsx).
    Retorna o caminho do arquivo gerado.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("openpyxl não instalado. Exportando como CSV...")
        return export_to_csv(leads, output_path)

    if output_path is None:
        output_path = Path("data/crm_export.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ─── Aba principal: Todos os Leads ─────────────────────────────────
    ws = wb.active
    ws.title = "CRM - Leads"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Status colors para formatação condicional
    status_colors = {
        "Prospect": "FFF2CC",       # amarelo claro
        "Qualificado": "D5E8D4",    # verde claro
        "Desqualificado": "F8CECC", # vermelho claro
        "Conexão Enviada": "DAE8FC", # azul claro
        "Conectado": "B6D7A8",      # verde médio
        "DM1 Enviada": "A9C4EB",    # azul médio
        "Respondeu": "93C47D",      # verde forte
        "Convertido": "6AA84F",     # verde escuro
    }

    # Escreve headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escreve dados
    for row_idx, lead in enumerate(leads, 2):
        row_data = _lead_to_row(lead)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Colorir linha pelo status
        status = lead["status"] if lead["status"] else ""
        if status in status_colors:
            fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type="solid",
            )
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Ajusta largura das colunas
    column_widths = {
        "A": 6, "B": 40, "C": 25, "D": 30, "E": 25,
        "F": 40, "G": 18, "H": 16, "I": 12, "J": 18,
        "K": 30, "L": 40, "M": 40, "N": 20, "O": 20,
        "P": 20, "Q": 20,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Congela a primeira linha (headers)
    ws.freeze_panes = "A2"

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    # ─── Aba de resumo ─────────────────────────────────────────────────
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.cell(row=1, column=1, value="Status").font = Font(bold=True, size=12)
    ws_resumo.cell(row=1, column=2, value="Quantidade").font = Font(bold=True, size=12)
    ws_resumo.cell(row=1, column=3, value="% do Total").font = Font(bold=True, size=12)

    # Conta por status
    from collections import Counter
    status_count = Counter(l["status"] for l in leads)
    total = len(leads)

    for row_idx, (status, count) in enumerate(
        sorted(status_count.items(), key=lambda x: -x[1]), 2
    ):
        ws_resumo.cell(row=row_idx, column=1, value=status)
        ws_resumo.cell(row=row_idx, column=2, value=count)
        pct = f"{count/total*100:.1f}%" if total > 0 else "0%"
        ws_resumo.cell(row=row_idx, column=3, value=pct)

        if status in status_colors:
            fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type="solid",
            )
            ws_resumo.cell(row=row_idx, column=1).fill = fill

    row_total = len(status_count) + 2
    ws_resumo.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws_resumo.cell(row=row_total, column=2, value=total).font = Font(bold=True)

    ws_resumo.column_dimensions["A"].width = 22
    ws_resumo.column_dimensions["B"].width = 14
    ws_resumo.column_dimensions["C"].width = 14

    # Salva
    wb.save(output_path)
    log.info(f"✓ Planilha exportada: {output_path}")
    return output_path


def export_to_csv(leads: list, output_path: Optional[Path] = None) -> Path:
    """Fallback: exporta como CSV caso openpyxl não esteja disponível."""
    import csv

    if output_path is None:
        output_path = Path("data/crm_export.csv")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for lead in leads:
            writer.writerow(_lead_to_row(lead))

    log.info(f"✓ CSV exportado: {output_path}")
    return output_path


def sync_lead_to_sheet(lead_data: dict, output_path: Optional[Path] = None) -> None:
    """
    Sincroniza um único lead com a planilha Excel.
    Se o lead já existir (por url_perfil), atualiza a linha.
    Se não existir, adiciona no final.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return  # silenciosamente ignora se openpyxl não estiver instalado

    if output_path is None:
        output_path = Path("data/crm_export.xlsx")

    if not output_path.exists():
        # Se a planilha não existe, não faz sync individual
        # (será criada no próximo export completo)
        return

    try:
        wb = load_workbook(output_path)
        ws = wb["CRM - Leads"]

        url = lead_data.get("url_perfil", "")
        if not url:
            return

        # Procura a linha existente pelo url_perfil (coluna B)
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == url:
                target_row = row
                break

        if target_row is None:
            # Adiciona nova linha
            target_row = ws.max_row + 1

        # Atualiza os campos que foram fornecidos
        col_map = {col: idx + 1 for idx, col in enumerate(COLUNAS)}
        for key, value in lead_data.items():
            if key in col_map:
                ws.cell(row=target_row, column=col_map[key], value=value or "")

        wb.save(output_path)
    except Exception as e:
        log.debug(f"Erro ao sincronizar lead na planilha: {e}")
